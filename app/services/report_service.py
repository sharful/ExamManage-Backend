"""
Report generation service.

Generates duty-list, room-schedule, and daily-schedule reports in PDF (ReportLab)
or Excel (openpyxl) format.
"""
import os
from datetime import date
from io import BytesIO
from typing import Literal, NamedTuple

_FONTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "fonts")
_COLLEGE_NAME = "কুমিল্লা সরকারি মহিলা কলেজ"

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.models.exam import Exam, ExamAssignment
from app.models.invigilator import Invigilator
from app.models.room import Room

ReportFormat = Literal["pdf", "excel"]


# ---------------------------------------------------------------------------
# Internal data fetch
# ---------------------------------------------------------------------------

class AssignmentRow(NamedTuple):
    exam_name: str
    time_slot: str          # "morning" / "evening"
    room_number: str
    seats: int
    head_name: str
    inv1_name: str
    inv2_name: str | None   # optional third invigilator


async def _fetch_rows(db: AsyncSession, target_date: date) -> list[AssignmentRow]:
    """Return one row per assignment (room+exam combo) for the given date."""
    HeadInv = aliased(Invigilator, name="head_inv")
    Inv1 = aliased(Invigilator, name="inv1")
    Inv2 = aliased(Invigilator, name="inv2")

    q = (
        select(
            Exam.exam_name,
            Exam.time_slot,
            Room.room_number,
            ExamAssignment.seats,
            HeadInv.name.label("head_name"),
            Inv1.name.label("inv1_name"),
            Inv2.name.label("inv2_name"),
        )
        .join(Exam, ExamAssignment.exam_id == Exam.id)
        .join(Room, ExamAssignment.room_id == Room.id)
        .join(HeadInv, ExamAssignment.head_invigilator_id == HeadInv.id)
        .join(Inv1, ExamAssignment.invigilator1_id == Inv1.id)
        .outerjoin(Inv2, ExamAssignment.invigilator2_id == Inv2.id)
        .where(Exam.exam_date == target_date)
        .order_by(Exam.time_slot, Room.room_number)
    )

    result = await db.execute(q)
    return [AssignmentRow(*row) for row in result.all()]


# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

def _pdf_doc(title: str, date_str: str) -> tuple:
    """Return (buffer, doc, styles) for a new PDF document."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    return buf, doc, styles


def _pdf_table(data: list[list], col_widths: list[float]):
    """Build a styled ReportLab Table."""
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                # Header row
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a3c5e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
                # Body rows
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("TOPPADDING", (0, 1), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                # Alternating row colours
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
                # Grid
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#aab4be")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ]
        )
    )
    return tbl


def _pdf_heading(text: str, styles):
    from reportlab.platypus import Paragraph

    style = styles["Heading1"]
    style.textColor = "#1a3c5e"
    return Paragraph(text, style)


def _pdf_subheading(text: str, styles):
    from reportlab.platypus import Paragraph

    return Paragraph(f"<b>{text}</b>", styles["Normal"])


def _has_bengali(text: str) -> bool:
    return any('\u0980' <= c <= '\u09FF' for c in text)


def _render_bangla_png(text: str, font_path: str, font_size_pt: int = 20, dpi: int = 300) -> bytes:
    """
    Shape Bengali (or mixed) text with HarfBuzz and render with FreeType so that
    complex-script conjuncts (e.g. ল্ল) form correctly — ReportLab's
    own TTFont renderer ignores OpenType GSUB substitution tables.
    """
    import uharfbuzz as hb
    import freetype
    from PIL import Image

    px_size = int(font_size_pt * dpi / 72)

    blob = hb.Blob.from_file_path(font_path)
    hb_face = hb.Face(blob)
    upem = hb_face.upem
    hb_font = hb.Font(hb_face)
    hb_font.scale = (upem, upem)

    hb_buf = hb.Buffer()
    hb_buf.add_str(text)
    hb_buf.guess_segment_properties()
    hb.shape(hb_font, hb_buf)

    infos = hb_buf.glyph_infos
    positions = hb_buf.glyph_positions
    px_per_unit = px_size / upem

    ft_face = freetype.Face(font_path)
    ft_face.set_pixel_sizes(0, px_size)
    ascender = ft_face.size.ascender >> 6
    descender = -(ft_face.size.descender >> 6)
    line_h = ascender + descender

    total_advance = int(sum(p.x_advance for p in positions) * px_per_unit)
    pad_x, pad_y = 10, 6
    img_w = total_advance + 2 * pad_x
    img_h = line_h + 2 * pad_y

    img = Image.new("RGB", (max(img_w, 1), max(img_h, 1)), (255, 255, 255))
    pixels = img.load()
    x_pen, baseline = pad_x, pad_y + ascender

    for info, pos in zip(infos, positions):
        glyph_id = info.codepoint
        x_adv = int(pos.x_advance * px_per_unit)
        x_off = int(pos.x_offset * px_per_unit)
        y_off = int(pos.y_offset * px_per_unit)
        try:
            ft_face.load_glyph(glyph_id, freetype.FT_LOAD_RENDER)
        except Exception:
            x_pen += x_adv
            continue
        bm = ft_face.glyph.bitmap
        left, top = ft_face.glyph.bitmap_left, ft_face.glyph.bitmap_top
        pitch = abs(bm.pitch)
        gx, gy = x_pen + x_off + left, baseline - top - y_off
        for row in range(bm.rows):
            for col in range(bm.width):
                alpha = bm.buffer[row * pitch + col]
                if alpha:
                    px_x, px_y = gx + col, gy + row
                    if 0 <= px_x < img_w and 0 <= px_y < img_h:
                        v = 255 - alpha
                        pixels[px_x, px_y] = (v, v, v)
        x_pen += x_adv

    out = BytesIO()
    img.save(out, format="PNG")
    return out.getvalue()


def _bangla_cell_image(text: str, font_path: str):
    """Return a ReportLab Image of Bengali text suitable for embedding in a table cell."""
    from reportlab.lib.units import cm
    from reportlab.platypus import Image as RLImage
    from PIL import Image as PILImage

    dpi = 150
    png = _render_bangla_png(text, font_path, font_size_pt=9, dpi=dpi)
    pil_img = PILImage.open(BytesIO(png))
    w_px, h_px = pil_img.size
    w_cm = w_px / dpi * 2.54
    h_cm = h_px / dpi * 2.54
    img = RLImage(BytesIO(png), width=w_cm * cm, height=h_cm * cm)
    img.hAlign = "LEFT"
    return img


def _prepare_pdf_data(data: list[list]) -> list[list]:
    """Replace any table cell containing Bengali text with a HarfBuzz-rendered PNG image."""
    font_path = os.path.join(_FONTS_DIR, "NikoshBAN.ttf")
    result = []
    for row in data:
        new_row = []
        for cell in row:
            if isinstance(cell, str) and _has_bengali(cell):
                new_row.append(_bangla_cell_image(cell, font_path))
            else:
                new_row.append(cell)
        result.append(new_row)
    return result


def _pdf_college_header():
    from reportlab.lib.units import cm
    from reportlab.platypus import Image as RLImage
    from PIL import Image as PILImage

    font_path = os.path.join(_FONTS_DIR, "NikoshBAN.ttf")
    dpi = 300
    png_bytes = _render_bangla_png(_COLLEGE_NAME, font_path, font_size_pt=20, dpi=dpi)

    pil_img = PILImage.open(BytesIO(png_bytes))
    w_px, h_px = pil_img.size
    w_cm = w_px / dpi * 2.54
    h_cm = h_px / dpi * 2.54

    rl_img = RLImage(BytesIO(png_bytes), width=w_cm * cm, height=h_cm * cm)
    rl_img.hAlign = "CENTER"
    return rl_img


def _excel_add_college_header(ws, num_cols: int) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    cell = ws["A1"]
    cell.value = _COLLEGE_NAME
    cell.font = Font(name="NikoshBAN", bold=True, size=20)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _excel_header_style():
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fill = PatternFill("solid", fgColor="1A3C5E")
    alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="AAB4BE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return font, fill, alignment, border


def _excel_body_style(row_idx: int):
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    font = Font(name="Calibri", size=10)
    fill_color = "FFFFFF" if row_idx % 2 == 0 else "F0F4F8"
    fill = PatternFill("solid", fgColor=fill_color)
    alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="AAB4BE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return font, fill, alignment, border


def _excel_write_headers(ws, headers: list[str]) -> None:
    h_font, h_fill, h_align, h_border = _excel_header_style()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = h_border
    ws.row_dimensions[1].height = 20


def _excel_write_row(ws, row_num: int, values: list) -> None:
    b_font, b_fill, b_align, b_border = _excel_body_style(row_num)
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = b_font
        cell.fill = b_fill
        cell.alignment = b_align
        cell.border = b_border


def _excel_autofit(ws, headers: list[str], all_rows: list[list]) -> None:
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row in all_rows:
            val = str(row[col_idx - 1] or "")
            if len(val) > max_len:
                max_len = len(val)
        # clamp width
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 2, 50)


def _excel_finalize(ws, num_rows: int, num_cols: int) -> None:
    from openpyxl.utils import get_column_letter

    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Duty list
# ---------------------------------------------------------------------------

async def generate_duty_list(
    db: AsyncSession,
    target_date: date,
    fmt: ReportFormat,
) -> tuple[bytes, str]:
    """
    Invigilator-centric duty list: one row per invigilator role per assignment.
    Returns (file_bytes, filename).
    """
    rows = await _fetch_rows(db, target_date)
    date_str = target_date.strftime("%d %B %Y")

    # Flatten to per-invigilator rows
    flat: list[list] = []
    for r in rows:
        slot = r.time_slot.value.capitalize()
        flat.append([r.head_name, "Head Invigilator", r.room_number, r.exam_name, slot])
        flat.append([r.inv1_name, "Invigilator", r.room_number, r.exam_name, slot])
        if r.inv2_name:
            flat.append([r.inv2_name, "Invigilator", r.room_number, r.exam_name, slot])

    headers = ["Invigilator Name", "Role", "Room", "Exam", "Time Slot"]

    if fmt == "pdf":
        return _duty_list_pdf(headers, flat, date_str), f"duty_list_{target_date}.pdf"
    return _duty_list_excel(headers, flat, date_str, target_date), f"duty_list_{target_date}.xlsx"


def _duty_list_pdf(headers: list[str], flat: list[list], date_str: str) -> bytes:
    from reportlab.lib.units import cm
    from reportlab.platypus import Spacer

    buf, doc, styles = _pdf_doc("Invigilator Duty List", date_str)
    story = [
        _pdf_college_header(),
        Spacer(1, 0.3 * cm),
        _pdf_heading(f"Invigilator Duty List — {date_str}", styles),
        Spacer(1, 0.4 * cm),
    ]

    col_widths = [6 * cm, 4 * cm, 3 * cm, 9 * cm, 3.5 * cm]
    data = _prepare_pdf_data([headers] + flat)
    story.append(_pdf_table(data, col_widths))

    doc.build(story)
    return buf.getvalue()


def _duty_list_excel(
    headers: list[str], flat: list[list], date_str: str, target_date: date
) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duty List"
    ws.sheet_properties.tabColor = "1A3C5E"

    _excel_add_college_header(ws, len(headers))

    # Title
    ws.merge_cells("A2:E2")
    from openpyxl.styles import Alignment, Font
    title_cell = ws["A2"]
    title_cell.value = f"Invigilator Duty List — {date_str}"
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="1A3C5E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    _excel_write_headers_offset(ws, headers, start_row=3)
    for i, row_data in enumerate(flat, start=4):
        _excel_write_row(ws, i, row_data)

    _excel_autofit(ws, headers, flat)
    _excel_finalize_offset(ws, len(flat), len(headers), header_row=3)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Room schedule
# ---------------------------------------------------------------------------

async def generate_room_schedule(
    db: AsyncSession,
    target_date: date,
    fmt: ReportFormat,
) -> tuple[bytes, str]:
    """
    Room-centric schedule: one row per assignment showing room, exam, invigilators.
    Returns (file_bytes, filename).
    """
    rows = await _fetch_rows(db, target_date)
    date_str = target_date.strftime("%d %B %Y")

    flat: list[list] = []
    for r in rows:
        slot = r.time_slot.value.capitalize()
        invigs = r.head_name + " (H), " + r.inv1_name
        if r.inv2_name:
            invigs += ", " + r.inv2_name
        flat.append([r.room_number, r.exam_name, slot, str(r.seats), invigs])

    headers = ["Room", "Exam", "Time Slot", "Seats", "Invigilators"]

    if fmt == "pdf":
        return _room_schedule_pdf(headers, flat, date_str), f"room_schedule_{target_date}.pdf"
    return _room_schedule_excel(headers, flat, date_str, target_date), f"room_schedule_{target_date}.xlsx"


def _room_schedule_pdf(headers: list[str], flat: list[list], date_str: str) -> bytes:
    from reportlab.lib.units import cm
    from reportlab.platypus import Spacer

    buf, doc, styles = _pdf_doc("Room Schedule", date_str)
    story = [
        _pdf_college_header(),
        Spacer(1, 0.3 * cm),
        _pdf_heading(f"Room Schedule — {date_str}", styles),
        Spacer(1, 0.4 * cm),
    ]

    col_widths = [3 * cm, 8 * cm, 3.5 * cm, 2.5 * cm, 11 * cm]
    data = _prepare_pdf_data([headers] + flat)
    story.append(_pdf_table(data, col_widths))

    doc.build(story)
    return buf.getvalue()


def _room_schedule_excel(
    headers: list[str], flat: list[list], date_str: str, target_date: date
) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Room Schedule"
    ws.sheet_properties.tabColor = "1A3C5E"

    _excel_add_college_header(ws, len(headers))

    ws.merge_cells("A2:E2")
    title_cell = ws["A2"]
    title_cell.value = f"Room Schedule — {date_str}"
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="1A3C5E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    _excel_write_headers_offset(ws, headers, start_row=3)
    for i, row_data in enumerate(flat, start=4):
        _excel_write_row(ws, i, row_data)

    _excel_autofit(ws, headers, flat)
    _excel_finalize_offset(ws, len(flat), len(headers), header_row=3)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Daily schedule (combined)
# ---------------------------------------------------------------------------

async def generate_daily_schedule(
    db: AsyncSession,
    target_date: date,
    fmt: ReportFormat,
) -> tuple[bytes, str]:
    """
    Full daily schedule: two sections — room schedule + duty list — in one file.
    Returns (file_bytes, filename).
    """
    rows = await _fetch_rows(db, target_date)
    date_str = target_date.strftime("%d %B %Y")

    room_rows: list[list] = []
    duty_rows: list[list] = []
    for r in rows:
        slot = r.time_slot.value.capitalize()
        invigs = r.head_name + " (H), " + r.inv1_name
        if r.inv2_name:
            invigs += ", " + r.inv2_name
        room_rows.append([r.room_number, r.exam_name, slot, str(r.seats), invigs])

        duty_rows.append([r.head_name, "Head Invigilator", r.room_number, r.exam_name, slot])
        duty_rows.append([r.inv1_name, "Invigilator", r.room_number, r.exam_name, slot])
        if r.inv2_name:
            duty_rows.append([r.inv2_name, "Invigilator", r.room_number, r.exam_name, slot])

    if fmt == "pdf":
        data = _daily_schedule_pdf(room_rows, duty_rows, date_str)
        return data, f"daily_schedule_{target_date}.pdf"
    data = _daily_schedule_excel(room_rows, duty_rows, date_str, target_date)
    return data, f"daily_schedule_{target_date}.xlsx"


def _daily_schedule_pdf(
    room_rows: list[list], duty_rows: list[list], date_str: str
) -> bytes:
    from reportlab.lib.units import cm
    from reportlab.platypus import PageBreak, Spacer

    buf, doc, styles = _pdf_doc("Daily Schedule", date_str)

    room_headers = ["Room", "Exam", "Time Slot", "Seats", "Invigilators"]
    duty_headers = ["Invigilator Name", "Role", "Room", "Exam", "Time Slot"]

    story = [
        _pdf_college_header(),
        Spacer(1, 0.3 * cm),
        _pdf_heading(f"Daily Exam Schedule — {date_str}", styles),
        Spacer(1, 0.3 * cm),
        _pdf_subheading("Room Schedule", styles),
        Spacer(1, 0.2 * cm),
        _pdf_table(_prepare_pdf_data([room_headers] + room_rows), [3 * cm, 8 * cm, 3.5 * cm, 2.5 * cm, 11 * cm]),
        Spacer(1, 0.6 * cm),
        _pdf_subheading("Invigilator Duty List", styles),
        Spacer(1, 0.2 * cm),
        _pdf_table(_prepare_pdf_data([duty_headers] + duty_rows), [6 * cm, 4 * cm, 3 * cm, 9 * cm, 3.5 * cm]),
    ]

    doc.build(story)
    return buf.getvalue()


def _daily_schedule_excel(
    room_rows: list[list], duty_rows: list[list], date_str: str, target_date: date
) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font

    wb = openpyxl.Workbook()

    # --- Sheet 1: Room Schedule ---
    ws1 = wb.active
    ws1.title = "Room Schedule"
    ws1.sheet_properties.tabColor = "1A3C5E"
    room_headers = ["Room", "Exam", "Time Slot", "Seats", "Invigilators"]
    _excel_add_college_header(ws1, len(room_headers))
    ws1.merge_cells("A2:E2")
    c = ws1["A2"]
    c.value = f"Room Schedule — {date_str}"
    c.font = Font(name="Calibri", bold=True, size=13, color="1A3C5E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 22
    _excel_write_headers_offset(ws1, room_headers, start_row=3)
    for i, row_data in enumerate(room_rows, start=4):
        _excel_write_row(ws1, i, row_data)
    _excel_autofit(ws1, room_headers, room_rows)
    _excel_finalize_offset(ws1, len(room_rows), len(room_headers), header_row=3)

    # --- Sheet 2: Duty List ---
    ws2 = wb.create_sheet("Duty List")
    ws2.sheet_properties.tabColor = "2E6DA4"
    duty_headers = ["Invigilator Name", "Role", "Room", "Exam", "Time Slot"]
    _excel_add_college_header(ws2, len(duty_headers))
    ws2.merge_cells("A2:E2")
    c2 = ws2["A2"]
    c2.value = f"Invigilator Duty List — {date_str}"
    c2.font = Font(name="Calibri", bold=True, size=13, color="1A3C5E")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 22
    _excel_write_headers_offset(ws2, duty_headers, start_row=3)
    for i, row_data in enumerate(duty_rows, start=4):
        _excel_write_row(ws2, i, row_data)
    _excel_autofit(ws2, duty_headers, duty_rows)
    _excel_finalize_offset(ws2, len(duty_rows), len(duty_headers), header_row=3)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Preview data (JSON)
# ---------------------------------------------------------------------------

async def get_preview_data(
    db: AsyncSession,
    report_type: str,
    target_date: date,
) -> dict:
    """Return JSON-serialisable preview data for the given report type."""
    rows = await _fetch_rows(db, target_date)

    if report_type == "duty-list":
        headers = ["Invigilator Name", "Role", "Room", "Exam", "Time Slot"]
        flat: list[list] = []
        for r in rows:
            slot = r.time_slot.value.capitalize()
            flat.append([r.head_name, "Head Invigilator", r.room_number, r.exam_name, slot])
            flat.append([r.inv1_name, "Invigilator", r.room_number, r.exam_name, slot])
            if r.inv2_name:
                flat.append([r.inv2_name, "Invigilator", r.room_number, r.exam_name, slot])
        return {"sections": [{"title": "Invigilator Duty List", "headers": headers, "rows": flat}]}

    if report_type == "room-schedule":
        headers = ["Room", "Exam", "Time Slot", "Seats", "Invigilators"]
        flat = []
        for r in rows:
            slot = r.time_slot.value.capitalize()
            invigs = r.head_name + " (H), " + r.inv1_name
            if r.inv2_name:
                invigs += ", " + r.inv2_name
            flat.append([r.room_number, r.exam_name, slot, str(r.seats), invigs])
        return {"sections": [{"title": "Room Schedule", "headers": headers, "rows": flat}]}

    # daily-schedule
    room_headers = ["Room", "Exam", "Time Slot", "Seats", "Invigilators"]
    duty_headers = ["Invigilator Name", "Role", "Room", "Exam", "Time Slot"]
    room_rows: list[list] = []
    duty_rows: list[list] = []
    for r in rows:
        slot = r.time_slot.value.capitalize()
        invigs = r.head_name + " (H), " + r.inv1_name
        if r.inv2_name:
            invigs += ", " + r.inv2_name
        room_rows.append([r.room_number, r.exam_name, slot, str(r.seats), invigs])
        duty_rows.append([r.head_name, "Head Invigilator", r.room_number, r.exam_name, slot])
        duty_rows.append([r.inv1_name, "Invigilator", r.room_number, r.exam_name, slot])
        if r.inv2_name:
            duty_rows.append([r.inv2_name, "Invigilator", r.room_number, r.exam_name, slot])
    return {
        "sections": [
            {"title": "Room Schedule", "headers": room_headers, "rows": room_rows},
            {"title": "Invigilator Duty List", "headers": duty_headers, "rows": duty_rows},
        ]
    }


# ---------------------------------------------------------------------------
# Shared Excel helpers (with title row offset)
# ---------------------------------------------------------------------------

def _excel_write_headers_offset(ws, headers: list[str], start_row: int) -> None:
    h_font, h_fill, h_align, h_border = _excel_header_style()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = h_border
    ws.row_dimensions[start_row].height = 20


def _excel_finalize_offset(ws, num_rows: int, num_cols: int, header_row: int) -> None:
    from openpyxl.utils import get_column_letter

    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(num_cols)}{header_row + num_rows}"
    )
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
